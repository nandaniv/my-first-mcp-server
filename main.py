from mcp.server.fastmcp import FastMCP
from typing import List
from openpyxl import load_workbook

EXCEL_FILE = "leaves.xlsx"


def get_workbook():
    return load_workbook(EXCEL_FILE)


# Create MCP serverc
mcp = FastMCP("LeaveManager")

# Tool: Check Leave Balance
@mcp.tool()
def get_leave_balance(employee_id: str) -> str:
    wb = get_workbook()

    sheet = wb["Employees"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        emp_id, balance = row

        if emp_id == employee_id:
            return f"{employee_id} has {balance} leave days remaining."

    return "Employee ID not found."

# Tool: Apply for Leave with specific dates
@mcp.tool()
def apply_leave(employee_id: str, leave_dates: list[str]) -> str:

    wb = get_workbook()

    emp_sheet = wb["Employees"]
    history_sheet = wb["LeaveHistory"]

    requested_days = len(leave_dates)

    for row in emp_sheet.iter_rows(min_row=2):

        if row[0].value == employee_id:

            balance = row[1].value

            if balance < requested_days:
                return (
                    f"Insufficient leave balance. "
                    f"Available: {balance}"
                )

            row[1].value = balance - requested_days

            for leave_date in leave_dates:
                history_sheet.append([
                    employee_id,
                    leave_date
                ])

            wb.save(EXCEL_FILE)

            return (
                f"Leave applied for {requested_days} day(s). "
                f"Remaining balance: {balance - requested_days}"
            )

    return "Employee ID not found."

# Resource: Leave history
@mcp.tool()
def get_leave_history(employee_id: str) -> str:
    wb = get_workbook()

    history_sheet = wb["LeaveHistory"]

    history = []

    for row in history_sheet.iter_rows(min_row=2, values_only=True):
        emp_id, leave_date = row

        if emp_id == employee_id:
            history.append(str(leave_date))

    if not history:
        return "No leave history found."

    return f"Leave history for {employee_id}: {', '.join(history)}"

# Resource: Greeting
@mcp.resource("greeting://{name}")
def get_greeting(name: str) -> str:
    """Get a personalized greeting"""
    return f"Hello, {name}! How can I assist you with leave management today?"

if __name__ == "__main__":
    mcp.run(transport="streamable-http")